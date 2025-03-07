from decimal import Decimal
from io import BytesIO
from django.db import IntegrityError, connection, connections
from django.contrib import messages
from django.http import HttpResponseRedirect
from django.urls import reverse
from django.http import JsonResponse
from django.views import View
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
import json
from django.contrib.staticfiles import finders
from django.core.exceptions import PermissionDenied
from django.http import HttpResponse
from django.urls import reverse_lazy
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.views.generic.edit import FormView
import openpyxl
from .models import ContaAPagar, ContaAReceber, FormasPagamento, FormasRecebimento, CategoriasFinanceiro, GerarParcela, CentroDeCusto, PlanoDeContas
from .forms import ContaAPagarForm, ContaAReceberForm, DateRangeForm, CategoriaForm, FormasPagamentoForm, FormasRecebimentoForm, GerarParcelasForm, CentroDeCustoForm, PlanoDeContasForm
from django.shortcuts import redirect, render
from django.db.models import Sum
from datetime import date, datetime, timedelta, timezone
from django.utils import timezone
from io import BytesIO
from django.http import HttpResponse
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle
from django.utils.dateparse import parse_date



class ContaAPagarListView(ListView):
    model = ContaAPagar
    template_name = 'conta_a_pagar_list.html'
    context_object_name = 'contas_a_pagar'
    paginate_by = 10

    def get_queryset(self):
       
        queryset = super().get_queryset()
        
        descricao = self.request.GET.get('descricao')
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')
        status = self.request.GET.get('status')

        start_date = parse_date(start_date) if start_date else None
        end_date = parse_date(end_date) if end_date else None

        if descricao:
            queryset = queryset.filter(descricao__icontains=descricao)
        
        if start_date and end_date:
            queryset = queryset.filter(data_emissao__range=(start_date, end_date))
        elif start_date:
            queryset = queryset.filter(data_emissao__gte=start_date)
        elif end_date:
            queryset = queryset.filter(data_emissao__lte=end_date)
        
        if status:
            queryset = queryset.filter(status_pagamento=status == 'True')

        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        contas = self.get_queryset()
        context['total_contas'] = contas.count()
        context['total_valor'] = contas.aggregate(total_valor=Sum('valor'))['total_valor'] or 0
        return context


class ContaAPagarDetailView(DetailView):
    model = ContaAPagar
    template_name = 'conta_a_pagar_detail.html'
    context_object_name = 'conta_a_pagar'




class ContaAPagarCreateView(CreateView):
    model = ContaAPagar
    template_name = 'conta_a_pagar_update.html'
    form_class = ContaAPagarForm
    success_url = reverse_lazy('conta_a_pagar_list')

    def form_valid(self, form):
        
        if not self.request.user.is_authenticated:
            return redirect('login')  # Ou redirecionar para a página desejada
        
        
        if not hasattr(self.request.user, 'empresa') or not self.request.user.empresa:
            raise PermissionDenied("Usuário não associado a uma empresa.")
        
        
        
        form.instance.empresa = self.request.user.empresa
        
        return super().form_valid(form)

class ContaAPagarUpdateView(UpdateView):
    model = ContaAPagar
    template_name = 'conta_a_pagar_update.html'
    form_class = ContaAPagarForm
    success_url = reverse_lazy('conta_a_pagar_list')

    def form_valid(self, form):
        
        if not self.request.user.is_authenticated:
            return redirect('login')  # Ou redirecionar para a página desejada
        
        
        if not hasattr(self.request.user, 'empresa') or not self.request.user.empresa:
            raise PermissionDenied("Usuário não associado a uma empresa.")
        
        conta = form.save(commit=False)
        
        if conta.status_pagamento:
            
            if not conta.data_pagamento:
                conta.data_pagamento = timezone.now().date()  
            conta.save()
        
        form.instance.empresa = self.request.user.empresa
        
        return super().form_valid(form)


class ContaAPagarDeleteView(DeleteView):
    model = ContaAPagar
    template_name = 'conta_a_pagar_delete.html'
    success_url = reverse_lazy('conta_a_pagar_list')

    def get_object(self):
        self.set_empresa()  
        return super().get_object()




class ContaAReceberListView(ListView):
    model = ContaAReceber
    template_name = 'conta_a_receber_list.html'
    context_object_name = 'contas_a_receber'
    paginate_by = 10

    def get_queryset(self):
       
        queryset = super().get_queryset()
        
        descricao = self.request.GET.get('descricao')
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')
        status = self.request.GET.get('status')

        start_date = parse_date(start_date) if start_date else None
        end_date = parse_date(end_date) if end_date else None

        if descricao:
            queryset = queryset.filter(descricao__icontains=descricao)
        
        if start_date and end_date:
            queryset = queryset.filter(data_emissao__range=(start_date, end_date))
        elif start_date:
            queryset = queryset.filter(data_emissao__gte=start_date)
        elif end_date:
            queryset = queryset.filter(data_emissao__lte=end_date)
        
        if status:
            queryset = queryset.filter(status_recebimento=status == 'True')

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        contas = self.get_queryset()
        context['total_contas'] = contas.count()
        context['total_valor'] = contas.aggregate(total_valor=Sum('valor'))['total_valor'] or 0
        return context


class ContaAReceberDetailView(DetailView):
    model = ContaAReceber
    template_name = 'conta_a_receber_detail.html'
    context_object_name = 'conta_a_receber'

    def get_object(self):

        return super().get_object()


class ContaAReceberCreateView(CreateView):
    model = ContaAReceber
    template_name = 'conta_a_receber_update.html'
    form_class = ContaAReceberForm
    success_url = reverse_lazy('conta_a_receber_list')

    def form_valid(self, form):
        
        if not self.request.user.is_authenticated:
            return redirect('login')  # Ou redirecionar para a página desejada
        
        
        if not hasattr(self.request.user, 'empresa') or not self.request.user.empresa:
            raise PermissionDenied("Usuário não associado a uma empresa.")
        
        form.instance.empresa = self.request.user.empresa
        
        return super().form_valid(form)


class ContaAReceberUpdateView(UpdateView):
    model = ContaAReceber
    template_name = 'conta_a_receber_update.html'
    form_class = ContaAReceberForm
    success_url = reverse_lazy('conta_a_receber_list')

    def form_valid(self, form):
        if not self.request.user.is_authenticated:
            return redirect('login')  # Ou redirecionar para a página desejada
        
        if not hasattr(self.request.user, 'empresa') or not self.request.user.empresa:
            raise PermissionDenied("Usuário não associado a uma empresa.")
        
        conta = form.save(commit=False)
        
        if conta.status_recebimento and not conta.data_recebimento:
            conta.data_recebimento = timezone.now().date()  # Corrigido para usar "=" e não "-"
            conta.save()
        
        form.instance.empresa = self.request.user.empresa
        
        return super().form_valid(form)



class ContaAReceberDeleteView(DeleteView):
    model = ContaAReceber
    template_name = 'conta_a_receber_delete.html'
    success_url = reverse_lazy('conta_a_receber_list')

    def get_object(self):
        
        return super().get_object()
    


def totaisapagar(request):
    contas = ContaAPagar.objects.all()
    total_contas = contas.count()
    total_valor = contas.aggregate(total_valor=Sum('valor'))['total_valor'] or 0
    
    context = {
        'total_contas': total_contas,
        'total_valor': total_valor
    }
    
    return render(request, 'conta_a_pagar_list.html', context)


def fluxo_caixa(request):
    start_date = date.today() - timedelta(days=30)
    end_date = date.today()
    
    if request.method == 'GET':
        form = DateRangeForm(request.GET)
        if form.is_valid():
            start_date = form.cleaned_data.get('start_date')
            end_date = form.cleaned_data.get('end_date')
    else:
        form = DateRangeForm(initial={'start_date': start_date, 'end_date': end_date})

    # Contas a receber e parcelas do tipo "a receber"
    entradas = (
        ContaAReceber.objects.filter(data_vencimento__range=[start_date, end_date]).aggregate(total=Sum('valor'))['total'] or 0
    ) + (
        GerarParcela.objects.filter(
            tipo="receber",
            vencimento_inicial__range=[start_date, end_date]
        ).aggregate(total=Sum('valor'))['total'] or 0
    )

 
    saidas = (
        ContaAPagar.objects.filter(data_vencimento__range=[start_date, end_date]).aggregate(total=Sum('valor'))['total'] or 0
    ) + (
        GerarParcela.objects.filter(
            tipo="pagar",
            vencimento_inicial__range=[start_date, end_date]
        ).aggregate(total=Sum('valor'))['total'] or 0
    )

    saldo_inicial = 0
    saldo_final = saldo_inicial + entradas - saidas

    context = {
        'form': form,
        'saldo_inicial': saldo_inicial,
        'entradas': entradas,
        'saidas': saidas,
        'saldo_final': saldo_final,
        'start_date': start_date,
        'end_date': end_date,
    }
    
    return render(request, 'fluxo_caixa.html', context)



def dash(request):
    start_date = date.today() - timedelta(days=30)
    end_date = date.today()

    if request.method == 'GET' and 'start_date' in request.GET and 'end_date' in request.GET:
        form = DateRangeForm(request.GET)
        if form.is_valid():
            start_date = form.cleaned_data.get('start_date')
            end_date = form.cleaned_data.get('end_date')
    else:
        form = DateRangeForm(initial={'start_date': start_date, 'end_date': end_date})

    # Contas a receber e parcelas do tipo "a receber"
    entradas = (
        ContaAReceber.objects.filter(data_vencimento__range=[start_date, end_date]).aggregate(total=Sum('valor'))['total'] or 0
    ) + (
        GerarParcela.objects.filter(
            tipo="receber",
            vencimento_inicial__range=[start_date, end_date]
        ).aggregate(total=Sum('valor'))['total'] or 0
    )

    # Contas a pagar e parcelas do tipo "a pagar"
    saidas = (
        ContaAPagar.objects.filter(data_vencimento__range=[start_date, end_date]).aggregate(total=Sum('valor'))['total'] or 0
    ) + (
        GerarParcela.objects.filter(
            tipo="pagar",
            vencimento_inicial__range=[start_date, end_date]
        ).aggregate(total=Sum('valor'))['total'] or 0
    )

    saldo_inicial = 0
    saldo = entradas - saidas

    context = {
        'form': form,
        'financeiro_dash': {
            'saldo_inicial': saldo_inicial,
            'entradas': entradas,
            'saidas': saidas,
            'saldo': saldo,
            'start_date': start_date,
            'end_date': end_date,
        }
    }

    return render(request, 'dash.html', context)




def exportar_contaapagar_excel(request):
    # Cria uma planilha do Excel
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Contas a Pagar'


    columns = ['Documento', 'Descrição', 'Valor', 'Emissão', 'Vencimento', 'Status', 'Categoria', 'Forma de Pagamento']
    worksheet.append(columns)

    contas = ContaAPagar.objects.all()  
    for conta in contas:
        worksheet.append([
            conta.documento,
            conta.descricao,
            conta.valor,
            conta.data_emissao.strftime('%d/%m/%Y') if conta.data_emissao else '',  
            conta.data_vencimento.strftime('%d/%m/%Y') if conta.data_vencimento else '',  
            conta.status_pagamento,
            str(conta.categorias),
            str(conta.forma_pagamento),
        ])
   
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=contas_a_pagar.xlsx'  # Nome do arquivo atualizado
    return response


def exportar_contaareceber_excel(request):
    
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Contas a Receber'

    # Define o cabeçalho
    columns = ['Documento', 'Descrição', 'Valor', 'Emissão', 'Vencimento', 'Status', 'Categoria', 'Forma de Recebimento']
    worksheet.append(columns)

    
    contas = ContaAReceber.objects.all() 
    for conta in contas:
        worksheet.append([
            conta.documento,
            conta.descricao,
            conta.valor,
            conta.data_emissao.strftime('%d/%m/%Y') if conta.data_emissao else '',
            conta.data_vencimento.strftime('%d/%m/%Y') if conta.data_vencimento else '',  
            conta.status_recebimento, 
            str(conta.categorias),
            str(conta.forma_recebimento)
        ])
   
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=contas_a_receber.xlsx'  
    return response



class FormaRecebimentoListView(ListView):
    model = FormasRecebimento
    template_name = 'formas_recebimento_list.html'
    context_object_name = 'formas_recebimento'
    paginate_by = 10

    def get_queryset(self):

        queryset = super().get_queryset()
        descricao = self.request.GET.get('descricao')
        
        if descricao:
            queryset = queryset.filter(descricao__icontains=descricao)
        
        return queryset


class FormaRecebimentoDetailView(DetailView):
    model = FormasRecebimento
    template_name = 'formas_recebimento_detail.html'
    context_object_name = 'forma_recebimento'

 

class FormaRecebimentoCreateView(CreateView):
    model = FormasRecebimento
    template_name = 'formas_recebimento_form.html'
    form_class = FormasRecebimentoForm
    success_url = reverse_lazy('formas_recebimento_list')

    def form_valid(self, form):
        
        if not self.request.user.is_authenticated:
            return redirect('login')  # Ou redirecionar para a página desejada
        
        
        if not hasattr(self.request.user, 'empresa') or not self.request.user.empresa:
            raise PermissionDenied("Usuário não associado a uma empresa.")
        
    
        form.instance.empresa = self.request.user.empresa
        form.instance.user = self.request.user
        
        return super().form_valid(form)


class FormaRecebimentoUpdateView(UpdateView):
    model = FormasRecebimento
    template_name = 'formas_recebimento_form.html'
    form_class = FormasRecebimentoForm
    success_url = reverse_lazy('forma_recebimento_list')

    def form_valid(self, form):
        
        if not self.request.user.is_authenticated:
            return redirect('login')  # Ou redirecionar para a página desejada
        
        
        if not hasattr(self.request.user, 'empresa') or not self.request.user.empresa:
            raise PermissionDenied("Usuário não associado a uma empresa.")
        

        form.instance.empresa = self.request.user.empresa
        
        return super().form_valid(form)


class FormaRecebimentoDeleteView(DeleteView):
    model = FormasRecebimento
    template_name = 'formas_recebimento_confirm_delete.html'
    success_url = reverse_lazy('forma_recebimento_list')



class FormaPagamentoListView(ListView):
    model = FormasPagamento
    template_name = 'formas_pagamento_list.html'
    context_object_name = 'formas_pagamento'
    paginate_by = 10

    def get_queryset(self):
       
        queryset = super().get_queryset()
        descricao = self.request.GET.get('descricao')
        
        if descricao:
            queryset = queryset.filter(descricao__icontains=descricao)
        
        return queryset


class FormaPagamentoDetailView(DetailView):
    model = FormasPagamento
    template_name = 'formas_pagamento_detail.html'
    context_object_name = 'forma_pagamento'




class FormaPagamentoCreateView(CreateView):
    model = FormasPagamento
    form_class = FormasPagamentoForm
    template_name = 'formas_pagamento_form.html'
    success_url = reverse_lazy('formas_pagamento_list')

    def form_valid(self, form):
        
        if not self.request.user.is_authenticated:
            return redirect('login')  # Ou redirecionar para a página desejada
        
        
        if not hasattr(self.request.user, 'empresa') or not self.request.user.empresa:
            raise PermissionDenied("Usuário não associado a uma empresa.")
        
    
        form.instance.empresa = self.request.user.empresa
        form.instance.user = self.request.user

        
        return super().form_valid(form)


class FormaPagamentoUpdateView(UpdateView):
    model = FormasPagamento
    form_class = FormasPagamentoForm
    template_name = 'formas_pagamento_form.html'
    success_url = reverse_lazy('formas_pagamento_list')

    def form_valid(self, form):
        
        if not self.request.user.is_authenticated:
            return redirect('login')  # Ou redirecionar para a página desejada
        
        
        if not hasattr(self.request.user, 'empresa') or not self.request.user.empresa:
            raise PermissionDenied("Usuário não associado a uma empresa.")
        

        form.instance.empresa = self.request.user.empresa
        
        return super().form_valid(form)


class FormaPagamentoDeleteView(DeleteView):
    model = FormasPagamento
    template_name = 'formas_pagamento_confirm_delete.html'
    success_url = reverse_lazy('formas_pagamento_list')


class CategoriaListView(ListView):
    model = CategoriasFinanceiro
    template_name = 'categoriasfinanceiro_lista.html'
    context_object_name = 'categorias'
    paginate_by = 10

    def get_queryset(self):
        
        queryset = super().get_queryset()
        descricao = self.request.GET.get('descricao')
        
        if descricao:
            queryset = queryset.filter(descricao__icontains=descricao)
        
        return queryset


class CategoriaDetailView(DetailView):
    model = CategoriasFinanceiro
    template_name = 'categoriasfinanceiro_detail.html'
    context_object_name = 'categoria'

    def get_object(self):

        return super().get_object()


class CategoriaCreateView(CreateView):
    model = CategoriasFinanceiro
    template_name = 'categoriasfinanceiro_form.html'
    form_class = CategoriaForm
    success_url = reverse_lazy('categoriasfinanceiro_list')

    def form_valid(self, form):
        
        if not self.request.user.is_authenticated:
            return redirect('login')  # Ou redirecionar para a página desejada
        
        
        if not hasattr(self.request.user, 'empresa') or not self.request.user.empresa:
            raise PermissionDenied("Usuário não associado a uma empresa.")
        
    
        form.instance.empresa = self.request.user.empresa
        form.instance.user = self.request.user
        
        return super().form_valid(form)


class CategoriaUpdateView(UpdateView):
    model = CategoriasFinanceiro
    template_name = 'categoriasfinanceiro_form.html'
    form_class = CategoriaForm
    success_url = reverse_lazy('categoriasfinanceiro_list')

    def form_valid(self, form):
        
        if not self.request.user.is_authenticated:
            return redirect('login')  # Ou redirecionar para a página desejada
        
        
        if not hasattr(self.request.user, 'empresa') or not self.request.user.empresa:
            raise PermissionDenied("Usuário não associado a uma empresa.")
        
    
        form.instance.empresa = self.request.user.empresa
        
        return super().form_valid(form)


class CategoriaDeleteView(DeleteView):
    model = CategoriasFinanceiro
    template_name = 'categoriasfinanceiro_confirm_delete.html'
    success_url = reverse_lazy('categoriasfinanceiro_list')





class GerarParcelasView(FormView):
    template_name = "gerar_parcelas.html"
    form_class = GerarParcelasForm
    success_url = "/parcelas/"

    def form_valid(self, form):
        # Configurar o banco de dados com base na empresa


        # Verificar se o usuário tem empresa associada
        if not self.request.user.empresa:
            messages.error(self.request, "Usuário não está associado a nenhuma empresa.")
            return self.form_invalid(form)

        # Obter os dados do formulário
        vencimento_inicial = form.cleaned_data["vencimento_inicial"]
        valor_total = form.cleaned_data["valor"]
        total_parcelas = form.cleaned_data["total_parcelas"]
        tipo = form.cleaned_data["tipo"]
        documento = form.cleaned_data["documento"]
        descricao = form.cleaned_data["descricao"]
        quitacao = form.cleaned_data["quitacao"]
        responsavel = form.cleaned_data["responsavel"]
        pagamento_total = form.cleaned_data["pagamento_total"]
        pagamento_parcial = form.cleaned_data["pagamento_parcial"]

        # Calcular o valor de cada parcela
        valor_parcela = round(valor_total / total_parcelas, 2)
        diferenca = round(valor_total - (valor_parcela * total_parcelas), 2)

        # Gerar parcelas
        data_atual = datetime.strptime(str(vencimento_inicial), "%Y-%m-%d")

        for i in range(1, total_parcelas + 1):
            data_vencimento = data_atual + timedelta(days=(i - 1) * 30)
            valor_atual = valor_parcela + (diferenca if i == total_parcelas else 0)

            GerarParcela.objects.create(
                valor=valor_atual,
                vencimento_inicial=data_vencimento,
                tipo=tipo,
                documento=documento,
                total_parcelas=total_parcelas,
                descricao=descricao,
                quitacao=quitacao,
                responsavel=responsavel,
                pagamento_total=pagamento_total,
                pagamento_parcial=pagamento_parcial,
                empresa=self.request.user.empresa,  
            )

        # Redirecionar para a lista de parcelas
        return HttpResponseRedirect(reverse("parcelas_geradas"))
    
    
@method_decorator(csrf_exempt, name='dispatch')
class AtualizarValorPagoView(View):
    def post(self, request, *args, **kwargs):
        try:
            # Obter os dados do corpo da requisição
            data = json.loads(request.body)
            parcela_id = data.get("parcela_id")
            valor_pago = data.get("valor_pago")

            # Verificar se o valor pago é um número válido
            if not isinstance(valor_pago, (int, float)):
                return JsonResponse({
                    "success": False,
                    "message": "O valor pago deve ser um número válido."
                }, status=400)

            # Buscar a parcela no banco de dados
            parcela = GerarParcela.objects.get(id=parcela_id)

            # Validação: impedir valores negativos
            if valor_pago < 0:
                return JsonResponse({
                    "success": False,
                    "message": "O valor pago não pode ser negativo."
                }, status=400)

            # Validação: impedir que o valor pago exceda o valor da parcela
            if valor_pago > parcela.valor:
                return JsonResponse({
                    "success": False,
                    "message": "O valor pago não pode ser maior que o valor total da parcela."
                }, status=400)

            # Atualizar os campos da parcela
            parcela.valor_pago = Decimal(valor_pago).quantize(Decimal('0.01'))  # Arredondar para 2 casas decimais
            parcela.pagamento_total = parcela.valor_pago == parcela.valor
            parcela.pagamento_parcial = 0 < parcela.valor_pago < parcela.valor
            parcela.save()

            # Retornar resposta de sucesso
            return JsonResponse({"success": True, "message": "Valor pago atualizado com sucesso."})
        except GerarParcela.DoesNotExist:
            # Caso a parcela não seja encontrada
            return JsonResponse({"success": False, "message": "Parcela não encontrada."}, status=404)
        except Exception as e:
            # Qualquer outro erro
            return JsonResponse({"success": False, "message": str(e)}, status=500)

    
from django.db.models import Sum, Count

class ParcelasListView(ListView):
    model = GerarParcela
    template_name = 'parcelas_geradas.html'
    context_object_name = 'parcelas_geradas'
    paginate_by = 10

    def get_queryset(self):

        queryset = super().get_queryset()
        descricao = self.request.GET.get('descricao')
        responsavel = self.request.GET.get('responsavel')
        data_inicial = self.request.GET.get('data_inicial')
        data_final = self.request.GET.get('data_final')

        if descricao:
            queryset = queryset.filter(descricao__icontains=descricao)
        
        if responsavel:
            queryset = queryset.filter(responsavel__nome__icontains=responsavel)

        if data_inicial and data_final:
            queryset = queryset.filter(vencimento_inicial__range=[data_inicial, data_final])

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Calcular o total e a contagem com base no queryset filtrado
        total_valores = self.get_queryset().aggregate(total=Sum('valor'))['total'] or 0
        total_parcelas = self.get_queryset().aggregate(quantidade=Count('id'))['quantidade'] or 0

        # Adicionar os valores ao contexto
        context['total_valores'] = total_valores
        context['total_parcelas'] = total_parcelas

        return context



class EditarParcelaView(UpdateView):
    model = GerarParcela
    form_class = GerarParcelasForm 
    template_name = 'editar_parcela.html'
    context_object_name = 'parcela'

    # Redirecionar para a lista de parcelas após a edição
    def get_success_url(self):
        return reverse_lazy('parcelas_list')
    


class ExcluirParcelaView(DeleteView):
    model = GerarParcela
    template_name = 'excluir_parcela.html'  
    context_object_name = 'parcela'
    success_url = reverse_lazy('parcelas_list')





class ListarCentrosDeCusto(ListView):
    model = CentroDeCusto
    template_name = 'cclistar.html'  
    context_object_name = 'centros'  



class CadastrarCentroDeCusto(CreateView):
    model = CentroDeCusto
    form_class = CentroDeCustoForm
    template_name = 'cccadastrar.html'  
    success_url = reverse_lazy('listar_centros_de_custo')  

    def form_valid(self, form):
        # Preenche o campo empresa com a empresa do usuário logado
        form.instance.empresa = self.request.user.empresa  # Ajuste isso conforme necessário
        return super().form_valid(form)

class EditarCentrosDeCustos(UpdateView):
    model = CentroDeCusto
    form_class = CentroDeCustoForm
    template_name =  'cccadastrar.html' 
    success_url = reverse_lazy('listar_centros_de_custo')  

    def form_valid(self, form):
        # Preenche o campo empresa com a empresa do usuário logado
        form.instance.empresa = self.request.user.empresa  # Ajuste isso conforme necessário
        return super().form_valid(form)
    

class DetalhesCentroDeCusto(DetailView):
    model = CentroDeCusto
    template_name = 'ccdetail.html'
    context_object_name = 'object'  # Para garantir que o contexto use 'object' como chave


    


class DeletarCentrosDeCustos(DeleteView):
    model = CentroDeCusto
    template_name =  'ccexcluir.html' 
    success_url = reverse_lazy('listar_centros_de_custo')  
    
    def form_valid(self, form):
        # Preenche o campo empresa com a empresa do usuário logado
        form.instance.empresa = self.request.user.empresa  # Ajuste isso conforme necessário
        return super().form_valid(form)
    


class ListarPlanoDeContas(ListView):
    model = PlanoDeContas
    template_name = 'planos_listar.html'  
    context_object_name = 'planos'  



class CadastrarPlanoDeContas(CreateView):
    model = PlanoDeContas
    form_class = PlanoDeContasForm
    template_name = 'planocadastrar.html'  
    success_url = reverse_lazy('planos_listar')  

    def form_valid(self, form):
        # Preenche o campo empresa com a empresa do usuário logado
        form.instance.empresa = self.request.user.empresa  # Ajuste isso conforme necessário
        return super().form_valid(form)

class EditarPlanoDeContas(UpdateView):
    model = PlanoDeContas
    form_class = PlanoDeContasForm
    template_name =  'planocadastrar.html' 
    success_url = reverse_lazy('planos_listar')  

    def form_valid(self, form):
        # Preenche o campo empresa com a empresa do usuário logado
        form.instance.empresa = self.request.user.empresa  # Ajuste isso conforme necessário
        return super().form_valid(form)
    

class DetalhesPlanoDeContas(DetailView):
    model = CentroDeCusto
    template_name = 'plano_detail.html'
    context_object_name = 'object'  


    


class DeletarPlanoDeContas(DeleteView):
    model = PlanoDeContas
    template_name =  'plano_excluir.html' 
    success_url = reverse_lazy('planos_listar')  
    
    def form_valid(self, form):
        # Preenche o campo empresa com a empresa do usuário logado
        form.instance.empresa = self.request.user.empresa  # Ajuste isso conforme necessário
        return super().form_valid(form)




def exportar_cc_excel(request):
    if not request.user.empresa:
        raise IntegrityError("Usuário não associado a uma empresa")
    

    
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'CentrosdeCustos'

    # Define o cabeçalho
    columns = ['Código', 'Nome', 'Pai', 'Tipo']
    worksheet.append(columns)

    # Adiciona os dados
    centros = CentroDeCusto.objects.all()  
    for centro in centros:
        worksheet.append([
            str(centro.codigo),
            str(centro.nome),
            str(centro.pai),
            str(centro.tipo)
            
        ])

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=cc.xlsx'
    return response




def gerar_tabela(centros):
   
   
    """Cria a tabela formatada para o PDF"""
    data = [["Código", "Nome", "Pai", "Tipo"]]  # Cabeçalhos
    
    for centro in centros:
        pai = centro.pai.nome if centro.pai else "N/A"
        data.append([str(centro.codigo), centro.nome, pai, centro.tipo])

    # Criação da tabela com os estilos 
    table = Table(data, colWidths=[80, 200, 120, 80])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
        ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
        ("GRID", (0, 0), (-1, -1), 1, colors.black),
    ]))
    return table

def exportar_cc_pdf(request):
    """Gera e retorna o relatório de Centros de Custo em PDF"""
    if not request.user.empresa:
        raise IntegrityError("Usuário não associado a uma empresa")

    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    
    # Carregar imagem
    image_path = finders.find('accounts/image/logo2.jpg')
    if image_path:
        pdf.drawImage(image_path, 50, height - 100, width=120, height=60)  # Ajustando posição

    # Título
    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawCentredString(width / 2, height - 50, "Relatório de Centros de Custo")

    # Obtendo os dados e gerando tabela
    centros = CentroDeCusto.objects.all()
    table = gerar_tabela(centros)

    # Centralizando a tabela
    table_width, table_height = table.wrap(0, 0)
    x_position = (width - table_width) / 2
    y_position = height - 180 - table_height  # Ajustar a posição vertical

    # Desenhar tabela no PDF
    table.drawOn(pdf, x_position, y_position)

    pdf.showPage()
    pdf.save()

    buffer.seek(0)
    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = 'inline; filename="cc.pdf"'
    return response